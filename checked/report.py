import csv
import math
import re
import sys
import os.path
import pdfkit
import numpy as np
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader, Template
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

dic_naming = {
    "name": "Название",
    "salary": "Оклад",
    "area_name": "Название региона",
    "published_at": "Дата публикации вакансии"
}
currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class UserInput:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.job_name = input('Введите название профессии: ')
        # self.choose = input('Введите данные для печати: ')

        # self.file_name = 'vacancies_by_year.csv'
        # self.job_name = 'Аналитик'
        # self.choose = 'Статистика'

        # self.choose_check()

    # def choose_check(self):
    #     if self.choose not in ['Вакансии', 'Статистика']:
    #         print('Некорректный аргумент. Введите: Вакансии или Статистика.')
    #         self.choose = input('Введите данные для печати: ')
    #         self.choose_check()


class DataSet:
    def __init__(self, st, user_input):
        self.file_name = user_input.file_name
        self.csv_headers, self.read_data = self.csv_reader()
        self.vacancies_objects = []
        self.csv_filter(st, user_input)

    def csv_reader(self):
        vacancies_file = open(self.file_name, encoding='utf_8_sig')
        vacancies_reader = csv.reader(vacancies_file)
        file_len = len(vacancies_file.readlines())
        if file_len == 0:
            print("Пустой файл")
            sys.exit()
        elif file_len == 1:
            print("Нет данных")
            sys.exit()
        vacancies_file.seek(0)
        return [vacancies_reader.__next__(), [x for x in vacancies_reader]]

    def csv_filter(self, st, user_input):
        for vacancy_data in self.read_data:
            if len(vacancy_data) == len(self.csv_headers) and vacancy_data.count('') == 0:
                self.vacancies_objects.append(Vacancy(vacancy_data, st, user_input))


class Vacancy:
    def __init__(self, vacancy_data, st, user_input):
        vacancy_data = [self.clean_value(s.split('\n')) for s in vacancy_data]
        i, c = 0, 0
        if len(vacancy_data) > 6:
            i, c = 5, 6
        self.name = vacancy_data[0]
        self.salary = Salary(vacancy_data[i + 1],
                             vacancy_data[i + 2],
                             vacancy_data[c + 3])
        self.area_name = vacancy_data[c + 4]
        self.published_at = self.date_format(vacancy_data[c + 5])

        st.add_data(self, user_input)

    @staticmethod
    def date_format(date):
        class dotdict(dict):  # Для получения значения через точку.
            """dot.notation access to dictionary attributes"""
            __getattr__ = dict.get
            __setattr__ = dict.__setitem__
            __delattr__ = dict.__delitem__

        return dotdict({'year': date[0:4]})  # Для статистики нужен только год.

    @staticmethod
    def clean_value(strings):
        html_tag_pattern = re.compile('<.*?>')
        result = []
        for string in strings:
            string = re.sub(html_tag_pattern, '', string)
            string = ' '.join(string.split())
            string = string.strip()
            result.append(string)
        if len(result) == 1:
            return result[0]
        return result


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = float(salary_from)
        self.salary_to = float(salary_to)
        self.salary_currency = salary_currency
        self.average_salary = math.floor((self.salary_from + self.salary_to) / 2) * \
                              currency_to_rub[self.salary_currency]


class Statistics:
    def __init__(self):
        self.year_to_salary = {}
        self.year_to_vac_num = {}
        self.year_to_salary_for_job = {}
        self.year_to_vac_num_for_job = {}
        self.area_to_salary = {}
        self.area_to_vac_num = {}
        self.sorted_area_salary = {}
        self.sorted_area_num = {}

    def add_data(self, vacancy, user_input):
        self.year_to_salary[vacancy.published_at.year] = self.year_to_salary.setdefault(vacancy.published_at.year, 0) + vacancy.salary.average_salary
        self.year_to_vac_num[vacancy.published_at.year] = self.year_to_vac_num.setdefault(vacancy.published_at.year, 0) + 1
        self.area_to_salary[vacancy.area_name] = self.area_to_salary.setdefault(vacancy.area_name, 0) + vacancy.salary.average_salary
        self.area_to_vac_num[vacancy.area_name] = self.area_to_vac_num.setdefault(vacancy.area_name, 0) + 1
        if user_input.job_name.lower() in vacancy.name.lower():
            self.year_to_salary_for_job[vacancy.published_at.year] = self.year_to_salary_for_job.setdefault(vacancy.published_at.year, 0) + vacancy.salary.average_salary
            self.year_to_vac_num_for_job[vacancy.published_at.year] = self.year_to_vac_num_for_job.setdefault(vacancy.published_at.year, 0) + 1
        else:
            self.year_to_salary_for_job[vacancy.published_at.year] = self.year_to_salary_for_job.setdefault(vacancy.published_at.year, 0)
            self.year_to_vac_num_for_job[vacancy.published_at.year] = self.year_to_vac_num_for_job.setdefault(vacancy.published_at.year, 0)

    def calc_average(self, data_set):
        self.year_to_salary = {k: self.average(k, v, self.year_to_vac_num) for k, v in self.year_to_salary.items()}
        self.area_to_salary = {k: self.average(k, v, self.area_to_vac_num) for k, v in self.area_to_salary.items()}
        self.year_to_salary_for_job = {k: self.average(k, v, self.year_to_vac_num_for_job) for k, v in self.year_to_salary_for_job.items()}

        area_salary_cut = {k: v for k, v in self.area_to_salary.items() if
                           self.area_to_vac_num[k] >= int(len(data_set.vacancies_objects) * 0.01)}
        self.sorted_area_salary = dict(sorted(area_salary_cut.items(), key=lambda i: i[1], reverse=True)[:10])

        area_num_cut = {k: round(v / len(data_set.vacancies_objects), 4) for k, v in self.area_to_vac_num.items() if
                        self.area_to_vac_num[k] >= int(len(data_set.vacancies_objects) * 0.01)}
        self.sorted_area_num = dict(sorted(area_num_cut.items(), key=lambda i: i[1], reverse=True)[:10])

    @staticmethod
    def average(year, num, dic):
        if year == 0 or num == 0:
            return 0
        return math.floor(num / dic[year])

    def print(self):
        print(f'Динамика уровня зарплат по годам: {self.year_to_salary}')
        print(f'Динамика количества вакансий по годам: {self.year_to_vac_num}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {self.year_to_salary_for_job}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {self.year_to_vac_num_for_job}')
        print(f'Уровень зарплат по городам (в порядке убывания): {self.sorted_area_salary}')
        print(f'Доля вакансий по городам (в порядке убывания): {self.sorted_area_num}')


class Report:
    def __init__(self, user_input):
        self.heads_by_year = ['Год', 'Средняя зарплата', f'Средняя зарплата - {user_input.job_name}',
                              'Количество вакансий', f'Количество вакансий - {user_input.job_name}']
        self.heads_by_area = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']

    def generate_pdf(self, st, user_input):
        self.generate_image(st, user_input)
        img_path = os.path.abspath('graph_t.png')
        table_html_template = '''
        <table>
            <tr>
                {% for head in heads %}
                {% if head == ' ' %}
                <th class="empty">{{ head }}</th>
                {% else %}                    
                <th>{{ head }}</th>
                {% endif %}
                {% endfor %}
            </tr>
            <tbody>
                {% for row in rows %}
                <tr>
                {% for cell in row %}
                {% if cell == ' ' %}
                <td class="empty">{{cell}}</td>
                {% else %}
                <td>{{cell}}</td>
                {% endif %}
                {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
        '''
        tm = Template(table_html_template)
        rows_1 = []
        for year, value in st.year_to_salary.items():
            rows_1.append([year, value, st.year_to_salary_for_job[year],
                           st.year_to_vac_num[year], st.year_to_vac_num_for_job[year]])
        table1 = tm.render(heads=self.heads_by_year, rows=rows_1)

        rows_2 = []
        for i in range(len(st.sorted_area_salary)):
            rows_2.append([list(st.sorted_area_salary.keys())[i], list(st.sorted_area_salary.values())[i], ' ',
                           list(st.sorted_area_num.keys())[i], f'{round(list(st.sorted_area_num.values())[i]*100, 2)}%'])
        table2 = tm.render(heads=self.heads_by_area, rows=rows_2)

        env = Environment(loader=FileSystemLoader('..'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'job_name': user_input.job_name, 'img_path': img_path, 'table1': table1, 'table2': table2})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def generate_image(self, st, user_input):
        fig = plt.figure()
        width = 0.4
        x_n = np.arange(len(st.year_to_salary.keys()))
        x_list1 = x_n - width / 2
        x_list2 = x_n + width / 2
        self.ax_1(st, user_input, fig, x_list1, x_list2, x_n, width)
        self.ax_2(st, user_input, fig, x_list1, x_list2, x_n, width)
        self.ax_3(st, fig)
        self.ax_4(st, fig)

        plt.tight_layout()
        plt.savefig('graph_t.png', dpi=300)

    @staticmethod
    def ax_1(st, user_input, fig, x_list1, x_list2, x_n, width):
        graf_1 = fig.add_subplot(221)
        graf_1.set_title('Уровень зарплат по городам')
        graf_1.bar(x_list1, st.year_to_salary.values(), width, label='средняя з/п')
        graf_1.bar(x_list2, st.year_to_salary_for_job.values(), width, label=f'з/п {user_input.job_name}')
        graf_1.set_xticks(x_n, st.year_to_salary.keys(), rotation='vertical')
        graf_1.tick_params(axis="both", labelsize=8)
        graf_1.legend(fontsize=8, loc='upper left')
        graf_1.grid(True, axis='y')

    @staticmethod
    def ax_2(st, user_input, fig, x_list1, x_list2, x_n, width):
        graf_2 = fig.add_subplot(222)
        graf_2.set_title('Количество вакансий по годам')
        graf_2.bar(x_list1, st.year_to_vac_num.values(), width, label='Количество вакансий')
        graf_2.bar(x_list2, st.year_to_vac_num_for_job.values(), width, label=f'Количество вакансий\n{user_input.job_name.lower()}')
        graf_2.set_xticks(x_n, st.year_to_salary.keys(), rotation='vertical')
        graf_2.tick_params(axis="both", labelsize=8)
        graf_2.legend(fontsize=8, loc='upper left')
        graf_2.grid(True, axis='y')

    @staticmethod
    def ax_3(st, fig):
        graf_3 = fig.add_subplot(223)
        graf_3.set_title('Уровень зарплат по городам')
        x_n = np.arange(len(st.sorted_area_salary.keys()))
        graf_3.barh(x_n, st.sorted_area_salary.values())
        graf_3.set_yticks(x_n, [key.replace(' ', '\n').replace('-', '-\n') for key in st.sorted_area_salary.keys()])
        graf_3.tick_params(axis='y', labelsize=6)
        graf_3.tick_params(axis="x", labelsize=8)
        graf_3.grid(True, axis='x')
        graf_3.invert_yaxis()

    @staticmethod
    def ax_4(st, fig):
        graf_4 = fig.add_subplot(224)
        graf_4.set_title('Доля вакансий по городам')
        data = [1 - sum(st.sorted_area_num.values())] + list(st.sorted_area_num.values())
        graf_4.pie(data, labels=['Другие'] + list(st.sorted_area_num.keys()), textprops={'fontsize': 6})
        graf_4.axis('equal')

    @staticmethod
    def cell_format(cell, align, is_area_page=False):
        thin = Side(border_style='thin', color='000000')

        if is_area_page and cell.column_letter == 'E':
            cell.number_format = FORMAT_PERCENTAGE_00

        if is_area_page and cell.column_letter == 'C':
            cell.border = Border(left=thin, right=thin)
        else:
            cell.border = Border(thin, thin, thin, thin)

        if cell.row != 1:
            cell.alignment = Alignment(horizontal=align)
        return Report.as_text(cell.value)

    @staticmethod
    def as_text(val):
        if val is None:
            return ""
        return str(val)

    def generate_excel(self, st, user_input):
        wb = Workbook()
        ws_year = wb.active
        ws_year.title = 'Статистика по годам'
        heads_by_year = ['Год', 'Средняя зарплата', f'Средняя зарплата - {user_input.job_name}', 'Количество вакансий', f'Количество вакансий - {user_input.job_name}']

        for i, head in enumerate(heads_by_year):
            ws_year.cell(1, i+1, head).font = Font(bold=True)
            ws_year.cell(1, i + 1).alignment = Alignment(horizontal='left')

        for year, value in st.year_to_salary.items():
            ws_year.append([year, value, st.year_to_salary_for_job[year], st.year_to_vac_num[year], st.year_to_vac_num_for_job[year]])

        for column in ws_year.columns:
            length = max(len(self.cell_format(cell, 'right')) for cell in column)
            ws_year.column_dimensions[column[0].column_letter].width = length + 2

        ws_area = wb.create_sheet('Статистика по городам')
        heads_by_area = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']
        for i, head in enumerate(heads_by_area):
            ws_area.cell(1, i + 1, head).font = Font(bold=True)
            ws_area.cell(1, i + 1).alignment = Alignment(horizontal='left')

        for i in range(len(st.sorted_area_salary)):
            ws_area.append([list(st.sorted_area_salary.keys())[i], list(st.sorted_area_salary.values())[i], ' ', list(st.sorted_area_num.keys())[i], list(st.sorted_area_num.values())[i]])

        for column in ws_area.columns:
            lengths = []
            for cell in column:
                if column[0].column_letter in ['B', 'E']:
                    self.cell_format(cell, 'right', True)
                else:
                    self.cell_format(cell, 'left', True)
                lengths.append(len(self.as_text(cell.value)))
            ws_area.column_dimensions[column[0].column_letter].width = max(lengths) + 2

        wb.save('report.xlsx')


def main():
    user_input = UserInput()
    st = Statistics()
    data_set = DataSet(st, user_input)
    st.calc_average(data_set)
    rep = Report(user_input)
    rep.generate_pdf(st, user_input)
